from students import Student
from openpyxl import load_workbook
import subprocess
import os


# Reads txt file where pre-defined input and expected output is stored
# Creates a dictionary for the values
def get_input_output(file):
    data = {}
    try:
        with open(file, "r") as f:
            tests = 0
            for line in f.readlines():
                script, ip, op = line.split("|")
                if script in data:
                    data[script] = data[script] + [(ip, op)]
                else:
                    data[script] = [(ip, op)]
                tests += 1
        return data, tests
    # Informs the user if their data file cannot be found
    except FileNotFoundError:
        print("I/O file does not exist.")
        exit(1)


# User must enter the complete file path
spreadsheet = input("Enter Excel file:")
io_file = input("Enter text file with i/o data:")

# Load Excel file where course roster and students' assignments are stored
try:
    # Ends program is the Excel is already open
    if os.path.isfile('~$' + spreadsheet):
        print("You must close the excel file.")
        exit(1)
    wb = load_workbook(spreadsheet)
    ws = wb.active
except FileNotFoundError:
    print("File does not exist.")
    exit(1)


# Student objects created will be stored in a dictionary
students = {}

# Input and output will be stored in a dictionary
# The weight of each assignment is calculated dynamically
io, programs = get_input_output(io_file)

# Creates Student objects using data from each row of the Excel file
for row in ws.values:
    first, last, id, folder, *data = row
    student = Student(first.strip(), last.strip(), id, folder)
    # Keeps header names out of the student dictionary
    if type(student.id) == int:
        students[student.id] = student

with open("output.txt", 'w') as f_out:
    # Iterates over student dictionary
    for id, student in students.items():
        assignments = os.listdir(student.scripts)  # List of programs
        path = os.path.abspath(student.scripts)  # File path for folder
        f_out.write(f'{student.first} {student.last}:\n')

        # Executes each one of the students' scripts
        correct = 0  # Counter for correctly made programs
        for script in assignments:
            for test in io[script]:
                # Name of each script must match the name written in the i/o txt file
                # Student will not receive credit if the file is named incorrectly
                ip, op = test

                # Runs command in the terminal
                result = subprocess.run(f"python {path}\\{script} {ip}", capture_output=True, shell=True, text=True)
                f_out.write(f'{script}: {ip} >>> {result.stdout}')

                # If the output matches the desired result, points are added to grade
                if op.strip() == result.stdout.strip():
                    correct += 1
                    f_out.write('Passed\n')
                # If an error occurred in the program, it will appear in the output file
                elif result.returncode != 0:
                    f_out.write(f'\nError: {result.stderr}\n')
                else:
                    f_out.write('Incorrect output\n')

        student.add_points(round(100 * correct / programs))
        f_out.write("\n")

# Adds new column that contains all grades to spreadsheet
i = 2
ws["E1"] = "Grades"
for s in students:
    ws["E" + str(i)] = students[s].grade
    i += 1

# Save the Excel file
wb.save("roster.xlsx")
